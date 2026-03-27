from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import re
from typing import Dict, Set, Tuple

from openpyxl import load_workbook


FILE_HEADER_CANDIDATES = {
    "file",
    "filename",
    "file_name",
    "excel",
    "excel_file",
    "workbook",
}
SHEET_HEADER_CANDIDATES = {"sheet", "sheet_name", "worksheet", "tab"}
COLUMN_HEADER_CANDIDATES = {
    "columns",
    "column",
    "column_name",
    "column_names",
    "mask_columns",
}
WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


RuleMap = Dict[str, Dict[str, Set[str]]]


def _normalize_header(header: str) -> str:
    return str(header).strip().lower().replace(" ", "_")


def _normalize_file_key(file_name: str) -> str:
    normalized = str(file_name).strip().lower()
    for ext in WORKBOOK_EXTENSIONS:
        if normalized.endswith(ext):
            return normalized[: -len(ext)]
    return normalized


def _split_columns(raw_value: object) -> Set[str]:
    if raw_value is None:
        return set()

    text = str(raw_value).strip()
    if not text:
        return set()

    parts = re.split(r"[,;|\n\r]+", text)
    return {part.strip() for part in parts if part.strip()}


def _find_header_column(headers: Dict[str, int], exact_candidates: Set[str], contains_text: str) -> int | None:
    exact_match = next((headers[h] for h in headers if h in exact_candidates), None)
    if exact_match:
        return exact_match
    return next((headers[h] for h in headers if contains_text in h), None)


def _find_multi_column_headers(headers: Dict[str, int]) -> list[int]:
    indexes: list[int] = []
    for header_key, col_idx in headers.items():
        if header_key.startswith("column"):
            indexes.append(col_idx)
    return sorted(indexes)


def load_rules(mapping_file: Path) -> RuleMap:
    workbook = load_workbook(mapping_file, data_only=True)
    sheet = workbook.active

    headers: Dict[str, int] = {}
    for index, cell in enumerate(sheet[1], start=1):
        if cell.value is None:
            continue
        headers[_normalize_header(str(cell.value))] = index

    file_col = _find_header_column(headers, FILE_HEADER_CANDIDATES, "file")
    sheet_col = _find_header_column(headers, SHEET_HEADER_CANDIDATES, "sheet")
    columns_col = _find_header_column(headers, COLUMN_HEADER_CANDIDATES, "column")
    multi_column_cols = _find_multi_column_headers(headers)

    if not file_col:
        raise ValueError(
            "Masking.xlsx must include a file name column. Supported examples: FileName, file_name"
        )

    # Locked format support: FileName + Sheet Name + Column1..ColumnN
    if not columns_col and not multi_column_cols:
        raise ValueError(
            "Masking.xlsx must include either a single columns field or Column1..ColumnN fields."
        )

    rules: RuleMap = defaultdict(lambda: defaultdict(set))
    for row in range(2, sheet.max_row + 1):
        file_name = sheet.cell(row=row, column=file_col).value
        if file_name is None:
            continue

        file_key = _normalize_file_key(str(file_name))
        if not file_key:
            continue

        sheet_key = "*"
        if sheet_col:
            sheet_value = sheet.cell(row=row, column=sheet_col).value
            if sheet_value is not None and str(sheet_value).strip():
                sheet_key = str(sheet_value).strip().lower()

        row_columns: Set[str] = set()
        if multi_column_cols:
            for col_idx in multi_column_cols:
                raw_col_name = sheet.cell(row=row, column=col_idx).value
                row_columns.update(_split_columns(raw_col_name))
        elif columns_col:
            raw_col_list = sheet.cell(row=row, column=columns_col).value
            row_columns.update(_split_columns(raw_col_list))

        if row_columns:
            rules[file_key][sheet_key].update(row_columns)

    return {k: dict(v) for k, v in rules.items()}


def resolve_rules_for_file(rules: RuleMap, input_file_name: str) -> Tuple[str | None, Dict[str, Set[str]]]:
    file_key = _normalize_file_key(input_file_name)
    if file_key in rules:
        return file_key, rules[file_key]

    # Fallback for locked templates that use base names while input files may contain suffixes.
    candidates = [
        rule_key
        for rule_key in rules
        if rule_key in file_key or file_key in rule_key
    ]
    if not candidates:
        return None, {}

    best_key = sorted(candidates, key=len, reverse=True)[0]
    return best_key, rules[best_key]
