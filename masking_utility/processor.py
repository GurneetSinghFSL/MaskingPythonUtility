from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, List, Set

from openpyxl import load_workbook

from .config import AppConfig
from .crypto import CryptoManager
from .rules import load_rules


@dataclass
class ProcessSummary:
    total_files: int = 0
    success_files: int = 0
    failed_files: int = 0
    total_cells_affected: int = 0
    errors: List[str] | None = None

    def __post_init__(self) -> None:
        if self.errors is None:
            self.errors = []


class WorkbookProcessor:
    def __init__(self, config: AppConfig, logger: logging.Logger) -> None:
        self.config = config
        self.logger = logger
        self.crypto = CryptoManager(config.pass_phrase)

    def run(
        self,
        mode: str,
        progress_callback: Callable[[int, str], None] | None = None,
    ) -> ProcessSummary:
        self.logger.info("run start | mode=%s", mode)

        summary = ProcessSummary()
        rules = load_rules(self.config.masking_rules_file)
        target_files = sorted(self.config.input_folder.glob("*.xlsx"))
        summary.total_files = len(target_files)

        output_dir = (
            self.config.output_folder_masked
            if mode == "mask"
            else self.config.output_folder_unmasked
        )
        output_dir.mkdir(parents=True, exist_ok=True)

        self.logger.info("total files discovered=%s", summary.total_files)
        for index, file_path in enumerate(target_files, start=1):
            self.logger.info("processing file start | file=%s", file_path)
            try:
                cell_count = self._process_file(file_path, output_dir, mode, rules)
                summary.total_cells_affected += cell_count
                summary.success_files += 1
                self.logger.info(
                    "processing file end | file=%s | cells_affected=%s",
                    file_path,
                    cell_count,
                )
            except Exception as exc:
                summary.failed_files += 1
                err = f"{file_path.name}: {exc}"
                summary.errors.append(err)
                self.logger.exception("processing file failed | file=%s", file_path)

            if progress_callback:
                progress_callback(index, file_path.name)

        self.logger.info(
            "run end | mode=%s | total=%s | success=%s | failed=%s | cells=%s",
            mode,
            summary.total_files,
            summary.success_files,
            summary.failed_files,
            summary.total_cells_affected,
        )
        return summary

    def _process_file(
        self,
        source_file: Path,
        output_dir: Path,
        mode: str,
        rules: Dict[str, Set[str]],
    ) -> int:
        self.logger.debug("_process_file start | file=%s", source_file)
        workbook = load_workbook(source_file)
        file_rule_columns = rules.get(source_file.name.lower(), set())
        total_affected = 0

        if not file_rule_columns:
            self.logger.info("no matching rule found; file copied with no changes | file=%s", source_file)

        for sheet in workbook.worksheets:
            total_affected += self._process_sheet(sheet, file_rule_columns, mode)

        output_path = output_dir / source_file.name
        workbook.save(output_path)
        self.logger.debug("_process_file end | file=%s | output=%s", source_file, output_path)
        return total_affected

    def _process_sheet(self, sheet, target_columns: Set[str], mode: str) -> int:
        self.logger.debug("_process_sheet start | sheet=%s", sheet.title)
        header_map = {}
        for col_idx in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row=1, column=col_idx).value
            if header_value is None:
                continue
            header_map[str(header_value).strip().lower()] = col_idx

        selected_indexes = [
            header_map[col.lower()] for col in target_columns if col.lower() in header_map
        ]

        affected = 0
        for col_idx in selected_indexes:
            for row_idx in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                original = str(cell.value)
                if mode == "mask":
                    new_value = self.crypto.mask(original)
                else:
                    new_value = self.crypto.unmask(original)

                if new_value != original:
                    cell.value = new_value
                    affected += 1

        self.logger.debug("_process_sheet end | sheet=%s | affected=%s", sheet.title, affected)
        return affected
